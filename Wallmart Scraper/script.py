import time
from tqdm import tqdm
import sys
import random
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import *
import re
import pandas as pd
from openpyxl import load_workbook
import math
from datetime import datetime,timedelta
import dateparser
from bs4 import BeautifulSoup as bs4




class WallMart:
	Url = ''
	#Total rows in excel file
	Rows = 0 
	## Defining options for chrome browser
	options = webdriver.ChromeOptions()
	options.add_argument("--ignore-certificate-errors")
	Browser = webdriver.Chrome(executable_path = "chromedriver",options = options)

	#Excel file declaration
	ExcelFile = pd.ExcelWriter('data.xlsx')
	#Creating constructor 
	def __init__(self,Url):
		self.Url = Url


	def ScrapeData(self,iH,cat):


		#Converting product div innerHTML to bueatifull soup 4
		bs = bs4(iH,"html.parser")



		#Extracting Type of the product if error then null
		try:
			Type = bs.find("span",{"class","flag-angle"}).text
		except:
			Type = ""




		#Extracting Type of the product
		try:
			Brand = bs.find("span",{"class","product-brand"}).text
		except:
			Brand = ""




		#Extracting Title of the product
		try:
			Title = bs.find("img")
			Title = Title['alt']
		except:
			Title = ""



		#Extracting All shipping methods
		try:
			Ship = bs.find("div",{"class","search-result-product-shipping-details gridview"})

			children = Ship.findChildren("div", recursive=False)

			#this piece of code will iterate through all the ship methods div and remove sold by div
			for c in range(len(children)):
				if "marketplace-sold-by" in children[c]['class']:
					children.pop(c)



			shipall = [shiptext.text for shiptext in children]

		except Exception as e:
			shipall =[""]
			


		#Exctracting sold by 
		try:
			sb = bs.find("div",{"class","marketplace-sold-by"}).text
		except:
			sb = ""
	




		#Extracting current oruce if the product
		try:
			Price = ""
			pricea = bs.find("span",{"class","price-main-block"})
			priceb= pricea.findAll("span",{"class","visuallyhidden"})
			if len(priceb) > 1:
				for price in priceb:
					Price = Price+price.text+" "
				Price = Price.rstrip()
				Price = Price.replace(' ','-')
			else:
				Price = priceb[0].text
		except Exception as e:
			Price = ""



		#Extracting price "Per product/ (13.6 ¢/ea)"
		try:
			op = bs.find("span", {"class","price-ppu-text"}).text
			op = op.rstrip(')').lstrip('(')
		except:
			op = ""

		#Extracting price sale price
		try:								
			op1 = bs.find("span",{"class","price display-inline-block arrange-fit price price-secondary"})
			op1 = op1.find("span",{"class","visuallyhidden"}).text
		except:
			op1 = ""


		currentdate = datetime.now()


		#if sale price is none 
		if op1 == "":
			#Appending data to dataframe
			df = pd.DataFrame({"TimeStamp":[currentdate],"Category": [cat],"Type": [Type],"Brand": [Brand],"Product": [Title],"Price": [Price],"Original Price": [op],"Sold By": [sb]})
			#Adding shipping method col dynamically
			for i in range(len(shipall)):
				seconddf = pd.DataFrame({"Delivery Type "+str(i):[shipall[i]]})
				df = df.join(seconddf)
		else:
			#Appending data to dataframe
			df = pd.DataFrame({"TimeStamp":[currentdate],"Category": [cat],"Type": [Type],"Brand": [Brand],"Product": [Title],"Price": [Price],"Original Price": [op1],"Sold By": [sb]})
			#Adding shipping method col dynamically
			for i in range(len(shipall)):
				seconddf = pd.DataFrame({"Delivery Type "+str(i+1):[shipall[i]]})
				df = df.join(seconddf)

		#If first entry in excel
		if self.Rows == 0:
			df.to_excel(self.ExcelFile,index=False,sheet_name='Data')
			self.Rows = self.ExcelFile.sheets['Data'].max_row
		else:
			df.to_excel(self.ExcelFile,index=False,sheet_name='Data',header=False,startrow=self.Rows)
			self.Rows = self.ExcelFile.sheets['Data'].max_row


		self.ExcelFile.save()




	def Main(self):
		self.Browser.get(self.Url)
		time.sleep(4)
		try:
			#Finding pagination Ul to get total number of pages
			pagination = self.Browser.find_elements_by_xpath("//ul[@class='paginator-list']//li")
			#last LI will always be the total page numbers
			TotalPages = int(pagination[-1].text)
		except:
			TotalPages = 1
		#Category text from a bread crump of search result       
		#There are two types of category innerHTML its going to try either of them 
		try:
			cat = self.Browser.find_element_by_xpath("//div[@class='search-breadcrumbs ']").text
		except:
			try:
				cat = self.Browser.find_element_by_xpath("//div[@class='merchant-module-wrapper container-full ResponsiveContainer']//div[@class='visualFacetTitle']").text
			except:
				cat = ""

		print("Total Pages found: "+str(TotalPages))

		for i in range(TotalPages):
			time.sleep(1)

			#Scraping total products present in the search result
			products = self.Browser.find_elements_by_xpath("//li[starts-with(@data-tl-id,'ProductTileGridView-')]")
			#Iterating through products
			for product in products:
				#Passing innerHTML of the product div and category text to the function which will extract all the data
				self.ScrapeData(product.get_attribute('innerHTML'),cat)

			print(str(i+1)+" page(s) done remaining "+str(TotalPages - (i+1)))
			
			try:
				#Clicking on next page button on each iteration if gets error then break the loop
				#Two types of next button
				try:
					NextButton = self.Browser.find_element_by_xpath("//button[@class='elc-icon paginator-hairline-btn paginator-btn paginator-btn-next']").click()
				except:
					NextButton = self.Browser.find_element_by_xpath("//button[@class='paginator-btn paginator-btn-next']").click()
			except:
				break
			time.sleep(4)
			








url1 = input("Enter Url : ")

#url1 = 'https://www.walmart.com/browse/clothing/men/5438_133197?povid=5438+%7C+2018-04-30+%7C+Fashion_LHN_Men_Clothing'
#url1= 'https://www.walmart.com/browse/household-essentials/kitchen/1115193_1071968'


a= WallMart(url1)
a.Main()